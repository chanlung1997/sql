import openpyxl as op
import json
wb = op.load_workbook(r'C:\\Users\Derek\\OneDrive - sagetech.com.hk\Desktop\\GIT\\testing\\datastan_batch1_20220109.xlsx')

sheetName = []
col1 = []
total={}
fName = 'C:\\Users\Derek\\OneDrive - sagetech.com.hk\Desktop\\GIT\\testing\\datastan_batch1_20221123.xlsx'

def readExcel():
    for filename in wb.sheetnames:
        sheet = wb[filename]
        line1 = str('f"'+'{PROJECT_ID}:'+sheet.cell(row=2, column=1).value+'.'+filename.lower()+'_raw"')
        sheetName.append(filename.lower())
        col1.append(line1)

def dictlist(sheetname,columnname):
    inValue = {'target_table_name': columnname, 'encoding':'utf-8', 'validation_method':'no_verify'}
    total[sheetname] = inValue

def main():
    readExcel()
    for x in range(0,len(sheetName)):
        dictlist(sheetName[x],col1[x])
    writejson(total)

def writejson(json_object):
    jsonString = json.dumps(json_object,indent=3)
    with open("data1.json", "w") as jsonFile:
        jsonFile.write(jsonString)
        jsonFile.close()
main()
        

