import openpyxl as op 

wb_datastan = op.load_workbook(r"C:\\Users\\Derek\\OneDrive - sagetech.com.hk\\Desktop\\GIT\\sql\\datastan_batch1_20230113_prod.xlsx")
wb_macro = op.load_workbook(r"C:\\Users\Derek\\OneDrive - sagetech.com.hk\Desktop\\GIT\\macro_rule_20221123.xlsx")
wb_unmatch = op.load_workbook(r"C:\\Users\\Derek\\OneDrive - sagetech.com.hk\\Desktop\\GIT\\unmatch column.xlsx")

# sheet_data = wb_datastan.sheetnames
sheet_unmatch_nm = wb_unmatch.sheetnames
# st_datastan = wb_datastan[sheet_unmatch_nm[1]]
macro_sheet = wb_macro["Model"]
# print(st_datastan)
# print(len(sheet_unmatch_nm))

for name in range(0,len(sheet_unmatch_nm)):
    target_table = sheet_unmatch_nm[name]
    # print(target_table)
    st_datastan = wb_datastan[sheet_unmatch_nm[name]]
    for i in range(1,st_datastan.max_row+1):
        for j in range(1,macro_sheet.max_row+1):
            if target_table.upper() == macro_sheet.cell(row=j, column=1).value.upper():
                if macro_sheet.cell(row=j, column=2).value == st_datastan.cell(row=i, column=4).value:
                    st_datastan.cell(row=i, column=9).value =="Y"
                    st_datastan.cell(row=i, column=10).value == macro_sheet.cell(row=j, column=4).value
                    st_datastan.cell(row=i, column=11).value == macro_sheet.cell(row=j, column=5).value
wb_datastan.save("datastan_batch1_20230113_prod.xlsx")