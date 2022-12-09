import openpyxl as op
import csv
wb = op.load_workbook(r"C:\\Users\\Derek\\OneDrive - sagetech.com.hk\\Desktop\\GIT\sql\\datastan_batch1_20221123.xlsx")

for sheet in wb:
    start_sen=[]
    
    for i in range(2,sheet.max_row):
        d = sheet.cell(row=i, column=4).value
        g = sheet.cell(row=i, column=7).value
        content=str(d)+" STRING "+"("+str(g)+")"+","
        start_sen.append(content)
    start = str("CREATE OR REPLACE TABLE `bqd_015_raw."+sheet.title+"_raw` (")
    start_sen.insert(0,start)
    end = str("__row_number STRING, \n__sys_code STRING, \n__file_name STRING, \n__batch_date TIMESTAMP, \n__load_date TIMESTAMP, \n__load_type STRING\n);")
    start_sen.append(end)
    path = "C:\\Users\\Derek\\OneDrive - sagetech.com.hk\\Desktop\\GIT\\sql\\"
    fileName = path+str(sheet.title)+'_raw'+'.sql'
    with open(fileName,'w',encoding='UTF8') as f:
        writer = csv.writer(f,delimiter='\n', quotechar='\"',quoting=csv.QUOTE_NONE,escapechar=' ')
        writer.writerow(start_sen)
        f.close()


    
