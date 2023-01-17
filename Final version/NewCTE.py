import openpyxl as op
import csv
wb = op.load_workbook(r"C:\Users\Derek\OneDrive - sagetech.com.hk\Desktop\GIT\sql\datastan_batch1_20230117_dev.xlsx")
# wb = op.load_workbook(r"C:\Users\Derek\OneDrive - sagetech.com.hk\Desktop\GIT\testing\datastan_batch1_20230113.xlsx")
env="dev"
for sheet in wb: 
    sheet = wb[sheet.title]
    casting= []
    macrofilename=[]
    columnselect = []
    validate_rule = []
    no_valid =[]
    aggrstr = ''
    n = 'NO'
    for i in range(2,sheet.max_row+1):
        d = sheet.cell(row=i, column=4)
        f = sheet.cell(row=i, column=6)
        g = sheet.cell(row=i, column=9)
        y = sheet.cell(row=i, column=8)
        j = sheet.cell(row=i, column=10)
        if d.value is None:
            break
        if y.value == 'Y':
            if d.value == "Title_Code": #title_code
                z = str('\t{{' + str(g.value) + "('" +"Gender_Code','"+ str(d.value)+"')" + '}}' + ',') #macro file name add-value
            elif d.value == "Phone_Area_Code": #phone
                z = str('\t{{' + str(g.value) + "('" +"Phone_Country_Code','"+ str(d.value)+"')" + '}}' + ',')
            elif d.value == "Doc_Nbr": #hkid
                z = str('\t{{' + str(g.value) + "('" +"Doc_Type_Code','"+ str(d.value)+"')" + '}}' + ',')
            else:
                z = str('\t{{' + str(g.value) + "('" + str(d.value)+"')" + '}}' + ',')
            validate_column = str("'__" + str(j.value) + "_valid_" + str(d.value) + "',")
            n = 'YES'
            aggrstr = aggrstr + validate_column
            macrofilename.append(z)
        tmp1 = str('\tSAFE_CAST('+str(d.value)+' AS '+ str(f.value) +') AS ' +str(d.value) + ',')
        casting.append(tmp1)
        tmp3 = "\t"+str(d.value) + ','
        columnselect.append(tmp3)
    tmp2 = f"WITH CTE_1 AS( \nSELECT\n\t*\nFROM gcp-prj-dat-bigquery-{env}-00.bqd_015_raw."+"tag_"f"{sheet.title.lower()}_raw\n), \n\nCTE_2 AS (\nSELECT"
    casting.insert(0,tmp2)
    if n =='YES':
        aggrstr = aggrstr[:-1]
        aggrstr = str('\t{{ full_valid_flag([') + aggrstr
        validate_rule.append(aggrstr + ']) }}}')
    casting[len(casting)-1] = str(casting[len(casting)-1][:-1])
    if n =='YES': 
        macrofilename[len(macrofilename)-1] = str(macrofilename[len(macrofilename)-1][:-1])
        validate_rule[len(validate_rule)-1] = str(validate_rule[len(validate_rule)-1][:-1])
        macrofilename.append('FROM CTE_2\n), \n\nCTE_4 AS (\nSELECT')
        validate_rule.append('FROM CTE_3)')
        validate_rule.append('\nSELECT * FROM CTE_4')
        
    else:
        no_valid.append("\t'' AS __full_load_valid \nFROM CTE_2\n)")
        validate_rule.append('\nSELECT * FROM CTE_3')
    casting.append('FROM CTE_1\n), \n\nCTE_3 AS (\n\tSELECT \n\t*,')
    if n=="YES":
        joinlist = casting + macrofilename + columnselect + validate_rule
    else:
        joinlist = casting + macrofilename + no_valid + validate_rule

    #export one file
    path = "C:\\Users\\Derek\\OneDrive - sagetech.com.hk\\Desktop\\GIT\\sql\\"
    fileName = path+"tag_"+sheet.title.lower()+'.sql'
    with open(fileName,'w',encoding='UTF8',newline='\n') as f:
        writer = csv.writer(f,delimiter='\n', quotechar='\"',quoting=csv.QUOTE_NONE,escapechar=' ')
        writer.writerow(joinlist)
        f.close()
exit()
