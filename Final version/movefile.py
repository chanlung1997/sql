import fnmatch
import shutil
import os 
import openpyxl as op
wb = op.load_workbook(r"C:\Users\Derek\OneDrive - sagetech.com.hk\Desktop\GIT\sql\datastan_batch1_20230117_dev.xlsx")
# wb2 = op.load_workbook(r"C:\Users\Derek\OneDrive - sagetech.com.hk\Desktop\GIT\sql\dbt_dags_list.xlsx")
#wb3 = op.load_workbook(r"C:\Users\Derek\OneDrive - sagetech.com.hk\Desktop\GIT\sql\UAT tables.xlsx")

sheet = wb.sheetnames
for file in os.listdir(r'C:\Users\Derek\OneDrive - sagetech.com.hk\Desktop\GIT\sql'):
    if file.startswith('tag'):
        if file[4:-4].upper() in sheet:
            shutil.move("C:\\Users\\Derek\\OneDrive - sagetech.com.hk\\Desktop\\GIT\\sql\\"+file, "C:\\Users\\Derek\\OneDrive - sagetech.com.hk\\Desktop\\GIT\\sql\\policytags\\"+file)
    else:
        pass

# additional cases

# sheet2 = wb2["Sheet1"]
# for i in range(1,62):
#     tableName = sheet2.cell(row=i, column=2).value
#     for file in os.listdir(r'C:\Users\Derek\OneDrive - sagetech.com.hk\Desktop\GIT\sql'):
#         if file.startswith('tag'):
#             if file[:-4] == tableName[4:]:
#                     shutil.move("C:\\Users\\Derek\\OneDrive - sagetech.com.hk\\Desktop\\GIT\\sql\\"+file, "C:\\Users\\Derek\\OneDrive - sagetech.com.hk\\Desktop\\GIT\\sql\\trypolicytags\\"+file)
#         else: 
#             pass

# sheet3 = wb3["Sheet1"]
# for i in range(1, sheet3.max_row+1):
#     tableName = sheet3.cell(row=i, column=1).value
#     for file in os.listdir(r'C:\Users\Derek\OneDrive - sagetech.com.hk\Desktop\GIT\sql\policytags'):
#         if file.startswith("tags"):
#             if file[:-4] == tableName:
#                 shutil.move("C:\\Users\\Derek\\OneDrive - sagetech.com.hk\\Desktop\\GIT\\sql\\policytags\\"+file, "C:\\Users\\Derek\\OneDrive - sagetech.com.hk\\Desktop\\GIT\\sql\\policytags\\New folder (2)\\"+file)
#             if file[:-4] == tableName[4:]:
#                 shutil.move("C:\\Users\\Derek\\OneDrive - sagetech.com.hk\\Desktop\\GIT\\sql\\policytags\\"+file, "C:\\Users\\Derek\\OneDrive - sagetech.com.hk\\Desktop\\GIT\\sql\\policytags\\New folder (2)\\"+file)
#         else:
#             pass