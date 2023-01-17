import openpyxl as op

wb = op.load_workbook(r"C:\Users\Derek\OneDrive - sagetech.com.hk\Desktop\GIT\sql\datastan_batch1_20230113_prod.xlsx")


# # for sheet in wb:
# #     for i in range(1,sheet.max_row):
# #         f = sheet.cell(row=i, column=6).value
# # #         g = sheet.cell(row=i, column=7).value
# #         if f =="DATE":
# #             sheet.cell(row=i, column=7).value= 8
# # wb.save("datastan_batch1_20221123.xlsx")

for sheet in wb:
    for i in range(2,sheet.max_row+1):
        e = sheet.cell(row=i, column=5).value 
        if e == "CF":
            sheet.cell(row=i, column=6).value = "CHARACTER"
            sheet.cell(row=i, column=7).value = "STRING"
        elif e == "CV":
            sheet.cell(row=i, column=6).value = "CHARACTER"
            sheet.cell(row=i, column=7).value = "STRING"  
        elif e == "D":       
            sheet.cell(row=i, column=6).value = "DECIMAL"
            sheet.cell(row=i, column=7).value = "NUMERIC"
        elif e == "DA":
            sheet.cell(row=i, column=6).value = "ANSIDATE"
            sheet.cell(row=i, column=7).value = "DATE"
        elif e == "I2":
            sheet.cell(row=i, column=6).value = "SMALLINT"
            sheet.cell(row=i, column=7).value = "INT64"
        elif e == "I":     
            sheet.cell(row=i, column=6).value = "INTEGER"
            sheet.cell(row=i, column=7).value = "INT64"       
wb.save("datastan_batch1_20230113_prod.xlsx")
