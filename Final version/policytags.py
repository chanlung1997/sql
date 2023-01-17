import openpyxl as op 
wb = op.load_workbook(r"C:\Users\Derek\OneDrive - sagetech.com.hk\Desktop\GIT\sql\CRM_Masking_List_v1.2 (002).xlsx")

sheet = wb["CRM"]
tableName = []
for k in range(1,sheet.max_row):
    c = sheet.cell(row = k, column=3)
    tableName.append(c.value)
tableName = list( dict.fromkeys(tableName) )
for l in tableName:
    top = "version: 2"+"\nmodels:"+"\n  - name: "+"tag_"+l.lower()+"\n    schema: bqd_995_marts"+"\n    config:"+"\n      alias: "+l.upper()+"\n    columns:"
    # top = "version: 2"+"\nmodels:"+"\n  - name: "+"tag_"+l.lower()+"\n    schema: bqd_995_marts"+"\n    config:"+"\n      alias: "+l.upper()
    fileName = "tag_"+l.lower()+'.yml'
    with open(fileName,'a')as file:
        file.writelines(top)
        file.close()
    for j in range(1, sheet.max_row):
        if l == sheet.cell(row=j, column=3).value:
            sens =str("\n    - name: "+sheet.cell(row=j,column=4).value+"\n      policy_tags:"+'\n        - '+'\'{{ var("DEFAULT_PT_TAXONOMY") ~ var("DEFAULT_PT_'+sheet.cell(row=j,column=6).value.upper()+'") }}\'')
            with open(fileName,'a')as file:
                file.writelines(sens)
                file.close()                       
        else:
            pass

