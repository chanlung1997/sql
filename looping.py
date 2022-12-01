import openpyxl as op
import csv
wb = op.load_workbook(r"C:\Users\Derek\OneDrive - sagetech.com.hk\Desktop\GIT\testing\datastan_batch1_20221123.xlsx")

for sheet in wb: 
    sheet = wb[sheet.title]
    casting= []
    macrofilename=[]
    columnselect = []
    validate_rule = []
    aggrstr = ''
    n = 'NO'
    for i in range(1,sheet.max_row):
        d = sheet.cell(row=i, column=4)
        f = sheet.cell(row=i, column=6)
        g = sheet.cell(row=i, column=9)
        y = sheet.cell(row=i, column=8)
        j = sheet.cell(row=i, column=10)
        if d.value is None:
            break
        if y.value == 'Y':
            z = str('\t{{' + str(g.value) + "('" + str(d.value)+"')" + '}}' + ',') #macro file name add-value
            validate_column = str("'__" + str(j.value) + "_valid_" + str(d.value) + "',")
            n = 'YES'
            aggrstr = aggrstr + validate_column
            macrofilename.append(z)
    
        tmp1 = str('\tCAST('+str(d.value)+' AS '+ str(f.value) +') AS ' +str(d.value) + ',')
        casting.append(tmp1)
        tmp3 = str(str(d.value) + ',')
        if n == 'YES':
            columnselect.append(tmp3)
    tmp2 = str('WITH CTE_1 AS( \nSELECT\n\t*\nFROM gcp-prj-dat-bigquery-env-00.bqd_015_raw.'+sheet.title+'_raw\n), \n\nCTE_2 AS (\nSELECT')
    casting.insert(0,tmp2)
    if n =='YES':
        aggrstr = aggrstr[:-1]
        aggrstr = str('\t{{ aggregate_validation_columns([') + aggrstr
        validate_rule.append(aggrstr + ']) }}}')
    casting[len(casting)-1] = str(casting[len(casting)-1][:-1])
    if n =='YES': 
        macrofilename[len(macrofilename)-1] = str(macrofilename[len(macrofilename)-1][:-1])
        validate_rule[len(validate_rule)-1] = str(validate_rule[len(validate_rule)-1][:-1])
        macrofilename.append('FROM CTE_2\n), \n\nCTE_4 AS (\nSELECT')
        validate_rule.append('FROM CTE_3)')
        validate_rule.append('\nSELECT * FROM CTE_4')
    else:
        columnselect.append("\t'' AS __full_load_valid \nFROM CTE_2\n)")
        validate_rule.append('\nSELECT * FROM CTE_3')
    casting.append('FROM CTE_1\n), \n\nCTE_3 AS (\n\tSELECT \n\t*,')
    joinlist = casting + macrofilename + columnselect + validate_rule

    #export one file
    path = "C:\\Users\\Derek\\OneDrive - sagetech.com.hk\\Desktop\\GIT\\sql\\"
    fileName = path+str(sheet.title)+'.sql'
    with open(fileName,'w',encoding='UTF8') as f:
        writer = csv.writer(f,delimiter='\n', quotechar='\"',quoting=csv.QUOTE_NONE,escapechar=' ')
        writer.writerow(joinlist)
        f.close()
