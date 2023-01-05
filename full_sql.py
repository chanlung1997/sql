import openpyxl as op
wb = op.load_workbook(r"C:\Users\Derek\OneDrive - sagetech.com.hk\Desktop\GIT\Macro_rule_20221123.xlsx")

sheet = wb["Model"]
tableName = []
for i in range(2, sheet.max_row):
    a = sheet.cell(row=i, column=1).value
    tableName.append(a)
tableName = list(dict.fromkeys(tableName))
###CTE_3
for j in tableName:
    full_load_valid = ''
    full_load_valid2 = ''
    full_valid_total = []
    full_load_valid_rule = []
    full_load_valid_rule2 = []
    cte3 = f"CTE_3 AS (\n\tSELECT\n\t*,"
    fileName = f"tag_{j.lower()}.sql"
    with open(fileName, "a") as f: 
        f.writelines(cte3)
        f.close()
    for k in range(2,sheet.max_row+1):
        if j == sheet.cell(row=k, column=1).value:   
            a = sheet.cell(row=k, column= 1).value
            b = sheet.cell(row=k, column= 2).value
            d = sheet.cell(row=k, column= 4).value
            e = sheet.cell(row=k, column= 5).value
            valid = f"\n\t{{{{{d}('{b}')}}}},"
            valid2= f"\n\t{{{{{d}('{b}')}}}}\nFROM CTE_2\n),"
            if sheet.cell(row=k, column=1).value != sheet.cell(row=k+1, column=1).value:
                with open(fileName,"a") as f:
                    f.writelines(valid2)
                    f.close()
            else:
                with open(fileName,"a") as f:
                    f.writelines(valid)
                    f.close() 
        else:
            pass
###CTE_4
    for l in range(2, sheet.max_row+1):
        if j == sheet.cell(row=l, column=1).value:
            b = sheet.cell(row=l, column= 2).value
            e = sheet.cell(row=l, column= 5).value
            full_valid = f"__{e}_valid_{b}',"
            full_valid2 = f"__{e}_valid_{b}'"
            if sheet.cell(row=l, column=1).value != sheet.cell(row=l+1, column=1).value:
                full_valid_total.append(full_valid2)
                
            else:
                full_valid_total.append(full_valid)
    aggr_column = "".join(full_valid_total)  
    full_load_valid = str('\n{{ full_valid_flag([') + aggr_column + ']) }}\nFROM CTE_3)'
    with open(fileName,"a") as f:
        f.writelines(full_load_valid)
        f.close()


 





